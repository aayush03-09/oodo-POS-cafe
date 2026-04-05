import io
from datetime import datetime, timedelta, date
from django.db.models import Sum, Count, F, Q
from django.utils import timezone
from rest_framework import status
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from django.http import HttpResponse
from apps.orders.models import Order, OrderItem, Payment
from apps.orders.serializers import OrderSerializer
from apps.bookings.models import Booking

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def dashboard(request):
    today = date.today()

    orders_today = Order.objects.filter(created_at__date=today)
    paid_orders = orders_today.filter(payment_status='done')
    pending_orders = orders_today.filter(payment_status='pending').exclude(status='draft')

    total_profit = Payment.objects.filter(
        order__created_at__date=today, status='confirmed'
    ).aggregate(total=Sum('amount'))['total'] or 0

    # Category sales
    category_sales = OrderItem.objects.filter(
        order__created_at__date=today, order__payment_status='done'
    ).values(category=F('product__category__name')).annotate(total=Sum('subtotal')).order_by('-total')

    # Daily profit 7 days
    daily_profit = []
    for i in range(6, -1, -1):
        day = today - timedelta(days=i)
        day_total = Payment.objects.filter(order__created_at__date=day, status='confirmed').aggregate(total=Sum('amount'))['total'] or 0
        daily_profit.append({'date': day.strftime('%Y-%m-%d'), 'day': day.strftime('%a'), 'total': float(day_total)})

    # Booking stats
    bookings_today = Booking.objects.filter(booking_date=today).exclude(status='cancelled')
    start_30 = today - timedelta(days=30)
    booking_by_date = Booking.objects.filter(booking_date__gte=start_30, booking_date__lte=today).exclude(status='cancelled').values('booking_date').annotate(count=Count('id')).order_by('booking_date')
    booking_graph = []
    for i in range(29, -1, -1):
        day = today - timedelta(days=i)
        count = Booking.objects.filter(booking_date=day).exclude(status='cancelled').count()
        booking_graph.append({'date': day.strftime('%Y-%m-%d'), 'day': day.strftime('%d %b'), 'bookings': count})

    # Orders for manager (all today with status + payment_status)
    recent_orders = Order.objects.filter(
        created_at__date=today
    ).exclude(status='draft').select_related('table__floor', 'created_by').prefetch_related('items__product', 'payments__method').order_by('-created_at')[:20]

    return Response({
        'total_profit': float(total_profit),
        'total_orders': orders_today.exclude(status='draft').count(),
        'paid_orders': paid_orders.count(),
        'pending_orders': pending_orders.count(),
        'open_sessions': 0,
        'total_bookings_today': bookings_today.count(),
        'category_sales': list(category_sales),
        'daily_profit': daily_profit,
        'booking_by_date': list(booking_by_date),
        'booking_graph': list(booking_graph),
        'recent_orders': OrderSerializer(recent_orders, many=True).data,
    })

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def sales_report(request):
    orders = Order.objects.filter(payment_status='done').select_related('table', 'created_by').prefetch_related('items__product', 'payments__method')
    date_from = request.query_params.get('date_from')
    date_to = request.query_params.get('date_to')
    if date_from: orders = orders.filter(created_at__date__gte=date_from)
    if date_to: orders = orders.filter(created_at__date__lte=date_to)

    total_revenue = Payment.objects.filter(order__in=orders, status='confirmed').aggregate(total=Sum('amount'))['total'] or 0
    top_products = OrderItem.objects.filter(order__in=orders).values(name=F('product__name')).annotate(total_qty=Sum('quantity'), total_sales=Sum('subtotal')).order_by('-total_sales')[:5]

    return Response({
        'total_revenue': float(total_revenue),
        'total_orders': orders.count(),
        'avg_order_value': float(total_revenue / orders.count()) if orders.count() > 0 else 0,
        'top_products': list(top_products),
        'orders': OrderSerializer(orders[:50], many=True).data,
    })

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def export_pdf(request):
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    styles = getSampleStyleSheet()
    elements = [Paragraph('POS Cafe - Sales Report', styles['Title']), Spacer(1, 20)]
    orders = Order.objects.filter(payment_status='done').select_related('table', 'created_by').prefetch_related('payments')

    data = [['Order #', 'Table', 'Staff', 'Amount', 'Date']]
    for order in orders[:100]:
        total = sum(p.amount for p in order.payments.all())
        data.append([order.order_number, order.table.table_number if order.table else '-', order.created_by.username, f'Rs.{total}', order.created_at.strftime('%Y-%m-%d %H:%M')])

    if len(data) > 1:
        table = Table(data)
        table.setStyle(TableStyle([('BACKGROUND', (0,0), (-1,0), colors.HexColor('#0ea5e9')), ('TEXTCOLOR', (0,0), (-1,0), colors.white), ('ALIGN', (0,0), (-1,-1), 'CENTER'), ('GRID', (0,0), (-1,-1), 1, colors.black)]))
        elements.append(table)

    doc.build(elements)
    buffer.seek(0)
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="sales_report.pdf"'
    return response

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def export_xls(request):
    import openpyxl
    from openpyxl.styles import Font, PatternFill
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = 'Sales Report'
    headers = ['Order #', 'Table', 'Staff', 'Amount', 'Payment Method', 'Date']
    fill = PatternFill(start_color='0ea5e9', end_color='0ea5e9', fill_type='solid')
    font = Font(color='FFFFFF', bold=True)
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h); cell.fill = fill; cell.font = font
    orders = Order.objects.filter(payment_status='done').select_related('table', 'created_by').prefetch_related('payments__method')
    for row, order in enumerate(orders[:100], 2):
        payment = order.payments.first()
        total = sum(p.amount for p in order.payments.all())
        ws.cell(row=row, column=1, value=order.order_number); ws.cell(row=row, column=2, value=order.table.table_number if order.table else '-')
        ws.cell(row=row, column=3, value=order.created_by.username); ws.cell(row=row, column=4, value=float(total))
        ws.cell(row=row, column=5, value=payment.method.name if payment and payment.method else '-'); ws.cell(row=row, column=6, value=order.created_at.strftime('%Y-%m-%d %H:%M'))
    buffer = io.BytesIO(); wb.save(buffer); buffer.seek(0)
    response = HttpResponse(buffer, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="sales_report.xlsx"'
    return response
